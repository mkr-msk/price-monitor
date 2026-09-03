"""Формирование отчёта об изменениях в Excel (.xlsx)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")

GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RED = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")


def _style_sheet(ws, widths: list[int], rows_after_header: int) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{max(rows_after_header + 1, 2)}"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=len(widths)):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT


def _add_sheet(wb, title: str, headers: list[str], rows: list[list], widths: list[int],
               fill_row_with_url_first_cell: bool = False) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    if fill_row_with_url_first_cell:
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("http"):
                    cell.fill = RED if "removed" in title.lower() else GREEN
    _style_sheet(ws, widths, len(rows))


def write_report(diff, output: str | Path) -> Path:
    """Собрать книгу отчёта: Сводка, Новые, Исчезли, Изменение цен."""
    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Сводка")
    summary.append(["Показатель", "Количество"])
    summary.append(["Новые позиции", len(diff.added)])
    summary.append(["Исчезнувшие позиции", len(diff.removed)])
    summary.append(["Изменилась цена", len(diff.price_changes)])
    summary.append(["Всего изменений", diff.total])
    _style_sheet(summary, [30, 12], summary.max_row - 1)

    if diff.added:
        _add_sheet(wb, "Новые", ["Название", "Цена", "Валюта", "Ссылка"],
                   [[p["title"], p["price"], p["currency"], p["url"]] for p in diff.added],
                   [60, 12, 10, 80], fill_row_with_url_first_cell=True)
    if diff.removed:
        _add_sheet(wb, "Исчезли", ["Название", "Цена", "Валюта", "Ссылка"],
                   [[p["title"], p["price"], p["currency"], p["url"]] for p in diff.removed],
                   [60, 12, 10, 80], fill_row_with_url_first_cell=True)
    if diff.price_changes:
        _add_sheet(wb, "Изменение цен",
                   ["Название", "Было", "Стало", "Изменение", "Ссылка"],
                   [[c.title, c.old, c.new,
                     round(c.new - c.old, 2) if c.old is not None and c.new is not None else None,
                     c.url] for c in diff.price_changes],
                   [60, 12, 12, 12, 80])

    wb.save(path)
    return path
